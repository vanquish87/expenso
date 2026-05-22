"""Excel import service: seed/re-seed CSVs from the source workbook.

The xlsx has:
 - "Categories" sheet: col A = GROUP / MAIN CATEGORY (merged cells —
   appears once per group, then None for siblings), col B = SUB-CATEGORY.
 - One sheet per month like "Apr-2026" ... "Mar-2027" with columns
   DATE, CATEGORY, NARRATION, DEBIT, CREDIT (row 1 is a banner, row 2
   is the header, data starts row 3).
"""
from __future__ import annotations

import re
from datetime import date as _date, datetime, time, timedelta
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

from ..core.exceptions import ValidationError
from ..domain.interfaces import BudgetRepository, CategoryRepository, EntryRepository
from ..domain.models import Category, Entry

MONTH_SHEET = re.compile(r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)-\d{4}$", re.I)


class ImportService:
    def __init__(
        self,
        categories: CategoryRepository,
        entries: EntryRepository,
        budgets: BudgetRepository,
    ) -> None:
        self._categories = categories
        self._entries = entries
        self._budgets = budgets

    def reseed_from_xlsx(self, path: Path) -> dict:
        if not path.exists():
            raise ValidationError(f"source workbook not found: {path}")

        wb = load_workbook(filename=str(path), data_only=True, read_only=True)

        cats = self._read_categories(wb)
        self._categories.replace_all(cats)

        # Importer goes around CategoryService.create, so explicitly drop
        # the template-side icon override cache.
        from ..templating import invalidate_category_icon_cache
        invalidate_category_icon_cache()

        valid_names = {c.name for c in cats}
        entries = self._read_entries(wb, valid_names)
        self._entries.replace_all(entries)

        return {
            "categories": len(cats),
            "entries": len(entries),
            "sheets_scanned": [s for s in wb.sheetnames if MONTH_SHEET.match(s)],
        }

    @staticmethod
    def _read_categories(wb) -> List[Category]:
        if "Categories" not in wb.sheetnames:
            return []
        ws = wb["Categories"]
        out: List[Category] = []
        seen_names: set[str] = set()
        current_group: Optional[str] = None

        # row 1 = banner, row 2 = header, data starts row 3
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:
                continue
            group_cell, sub_cell = (row + (None, None))[:2]
            group_val = (str(group_cell).strip() if group_cell else "") or None
            sub_val = (str(sub_cell).strip() if sub_cell else "") or None

            if group_val:
                current_group = group_val
                if group_val not in seen_names:
                    out.append(Category(name=group_val, parent=None))
                    seen_names.add(group_val)

            if sub_val and current_group:
                if sub_val not in seen_names:
                    out.append(Category(name=sub_val, parent=current_group))
                    seen_names.add(sub_val)

        return out

    @staticmethod
    def _read_entries(wb, valid_names: set) -> List[Entry]:
        """Walk every month sheet in workbook order, row-by-row.

        Each entry gets a synthetic timestamp = ``date_at_midnight + N minutes``
        where N is the per-date counter incremented in walk order. This
        preserves xlsx top-to-bottom ordering inside any single date when
        the ledger later sorts by (date, timestamp).
        """
        out: List[Entry] = []
        per_date_count: Dict[_date, int] = {}
        for sheet_name in wb.sheetnames:
            if not MONTH_SHEET.match(sheet_name):
                continue
            ws = wb[sheet_name]
            # row 1 = banner, row 2 = header, data starts row 3
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i < 2:
                    continue
                row = list(row) + [None] * (5 - len(row))
                date_v, cat_v, narr_v, debit_v, credit_v = row[:5]
                if not date_v or not cat_v:
                    continue
                if isinstance(date_v, datetime):
                    d = date_v.date()
                elif isinstance(date_v, _date):
                    d = date_v
                else:
                    try:
                        d = datetime.fromisoformat(str(date_v)).date()
                    except ValueError:
                        continue
                cat = str(cat_v).strip()
                if not cat or cat not in valid_names:
                    continue
                narr = (str(narr_v).strip() if narr_v else "")
                try:
                    debit = float(debit_v) if debit_v not in (None, "") else 0.0
                except (TypeError, ValueError):
                    debit = 0.0
                try:
                    credit = float(credit_v) if credit_v not in (None, "") else 0.0
                except (TypeError, ValueError):
                    credit = 0.0
                if debit == 0.0 and credit == 0.0:
                    continue
                if debit > 0 and credit > 0:
                    # ambiguous row — pick the larger side, leave the other 0
                    if debit >= credit:
                        credit = 0.0
                    else:
                        debit = 0.0
                n = per_date_count.get(d, 0)
                per_date_count[d] = n + 1
                ts = datetime.combine(d, time(0, 0)) + timedelta(minutes=n)
                out.append(
                    Entry(
                        date=d,
                        timestamp=ts,
                        category=cat,
                        narration=narr,
                        debit=debit,
                        credit=credit,
                    )
                )
        return out
